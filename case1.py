import openpyxl  # 导入Excel库
import requests  # 导入requests库


# 读取测试用例函数
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿名字  文档名字
    sheet = wb[sheetname]  # 获取表单
    max_row = sheet.max_row  # 获取最大行数
    case_list = []  # 创建空列表 存放读取数据
    for i in range(2, max_row+1):
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,  # 获取用例id
            url=sheet.cell(row=i, column=5).value,  # 获取url
            data=sheet.cell(row=i, column=6).value,  # 获取data
            expect=sheet.cell(row=i, column=7).value, )  # 定义字典存放一条Excel数据
        case_list.append(dict1)  # 存放Excel数据到列表
    return case_list   # 返回测试用例列表


# 写入Excel文件函数
def write_result(filename, sheetname, row, column, data):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿名字  文档名字
    sheet = wb[sheetname]  # 获取表单
    cell = sheet.cell(row=row, column=column).value = data    # 写入表单数据
    wb.save(filename)   # 保存,关闭文档


# 注册函数
def register(url, data):
    headers = {
        'X-Lemonban-Media-Type': 'lemonban.v2',
        'Content-Type': 'application/json'
    }  # 请求头
    requests1 = requests.post(url=url, headers=headers, json=data)  # 发送get请求并定义requests1变量存放响应信息
    return requests1.json()   # 返回json格式的响应内容


# 登录函数
def api_fun(url, data):
    headers = {
        'X-Lemonban-Media-Type': 'lemonban.v2'
    }  # 请求头
    res1 = requests.post(url=url, headers=headers, json=data)  # 发送get请求并定义requests1变量存放响应信息
    return res1.json()  # 返回响应内容


# 执行测试用例封装函数
def test(excel, sheet, function):
    cases = read_data(excel, sheet)  # 调用读取测试用例函数并保存返回测试列表
    for case in cases:   # 循环取出测试用例执行
        id = case.get('case_id')  # 获取用例id
        url = case.get('url')   # 获取url
        data = eval(case.get('data'))  # 获取data  eval()去掉字符串引号
        expect = eval(case.get('expect'))  # 获取预期结果 eval()去掉字符串引号
        msg = expect.get('msg')   # 获取预期结果中的msg
        register1 = function(url=url, data=data)   # 传参调用注册请求函数 并用变量接收
        register_msg = register1.get('msg')  # 获取实际结果中的msg
        print('预期结果中的msg：{}'.format(msg))
        print('实际结果中的msg：{}'.format(register_msg))
        if register_msg == msg:  # 判断预期结果是否与实际结果一致，并在文件中写入测试结果
            print("第{}条测试用例执行通过".format(id))
            write_result(excel, sheet, id+1, 8, 'Passed')   # 调用写入excel文件函数 写入passed
        else:
            print("第{}条测试用例执行未通过".format(id))
            write_result(excel, sheet, id+1, 8, 'Failed')  # 调用写入excel文件函数 写入Failed
        print('*'*15)


# 调用测试用例函数
test('test_case_api.xlsx', 'register', register)  # 执行注册测试用例
test('test_case_api.xlsx', 'login', api_fun)  # 执行登录测试用例
